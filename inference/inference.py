# -*- coding: utf-8 -*-
"""inference_cv.py — Cross-validation version of inference.py

Replaces the fixed pre-defined train/test splits with k-fold cross-validation
built on-the-fly from all available video keys in the dataset.  Every other
detail (model, evaluation metrics, xlsx export, parallel scan) is preserved.

Key changes vs. the original:
  * --n_folds  (default 5) controls the number of CV folds.
  * --cv_seed  (default 42) seeds the fold shuffle for reproducibility.
  * The split_file is still read, but only to discover ALL available keys;
    the train/test partition is then created by KFold instead of being read
    directly from the JSON.
  * All reporting functions are reused verbatim or with minimal adaptation.
"""

import torch
from os import listdir
import numpy as np
from os.path import join
import h5py
import json
import argparse
import re
import os
import logging
from concurrent.futures import ProcessPoolExecutor, as_completed
from sklearn.model_selection import KFold

from utils.utils import get_paths, setup_logging
from evaluation.evaluation_metrics import evaluate_summary
from model.layers.summarizer import xLSTM
from inference.generate_summary import generate_summary
from scipy.stats import kendalltau, spearmanr

setup_logging()


# ---------------------------------------------------------------------------
# Data loading  (unchanged)
# ---------------------------------------------------------------------------

def load_video_data(dataset, data_path, video):
    """Load video data from the dataset h5 file."""
    with h5py.File(data_path, "r") as hdf:
        frame_features = torch.Tensor(
            np.array(hdf[f"{video}/features"])
        ).view(-1, 1024)
        sb         = np.array(hdf[f"{video}/change_points"])
        video_name = None

        if dataset.lower() in ('summe', 'tvsum'):
            user_summary = np.array(hdf[f"{video}/user_summary"])
            n_frames     = np.array(hdf[f"{video}/n_frames"])
            positions    = np.array(hdf[f"{video}/picks"])
            if "video_name" in hdf[f"{video}"]:
                video_name = str(
                    np.array(hdf[f"{video}/video_name"]).astype(str, copy=False)
                )
        elif dataset.lower() == 'mrhisum':
            user_summary = np.array(hdf[f"{video}/gt_summary"])
            n_frames     = frame_features.shape[0]
            positions    = np.arange(n_frames, dtype=int)
        else:
            raise ValueError(f"Unsupported dataset: {dataset}")

    return frame_features, user_summary, sb, n_frames, positions, video_name


# ---------------------------------------------------------------------------
# Best-epoch selection helpers  (unchanged)
# ---------------------------------------------------------------------------

def _find_epoch_files(model_path):
    """Return all epoch-N.pkl files in model_path, sorted by epoch number."""
    files = [
        f for f in listdir(model_path)
        if re.match(r"epoch-\d+\.pkl", f)
    ]
    return sorted(files, key=lambda x: int(re.findall(r'\d+', x)[0]))


def _load_best_epoch_from_fscores(model_path):
    """Read pre-computed f_scores.txt produced by compute_fscores.py.

    Returns the best epoch number (0-indexed line = 0-indexed epoch).
    Returns None if the file does not exist.
    """
    fscores_path = join(model_path, 'f_scores.txt')
    if not os.path.exists(fscores_path):
        return None
    with open(fscores_path) as fp:
        content = fp.read().strip()
    try:
        scores = json.loads(content)
    except json.JSONDecodeError:
        scores = [float(x) for x in content.splitlines()]
    return int(np.argmax(scores))


# ---------------------------------------------------------------------------
# NEW: Cross-validation fold builder
# ---------------------------------------------------------------------------

def build_cv_folds(all_keys, n_folds=5, seed=42):
    """Partition *all_keys* into n_folds (train_keys, test_keys) pairs.

    Uses sklearn's KFold with a fixed random seed so the split is
    deterministic and reproducible.  The fold index plays the same role as
    ``split_id`` in the original code.

    Args:
        all_keys (list[str]): All video identifiers available in the dataset.
        n_folds  (int):       Number of CV folds (default 5).
        seed     (int):       Shuffle seed for reproducibility (default 42).

    Returns:
        list[dict]: Each element has keys ``"train_keys"`` and ``"test_keys"``.
    """
    keys_arr = np.array(all_keys)
    kf       = KFold(n_splits=n_folds, shuffle=True, random_state=seed)
    folds    = []
    for train_idx, test_idx in kf.split(keys_arr):
        folds.append({
            "train_keys": keys_arr[train_idx].tolist(),
            "test_keys":  keys_arr[test_idx].tolist(),
        })
    logging.info(
        f"Built {n_folds}-fold CV — "
        f"~{len(folds[0]['test_keys'])} test videos per fold "
        f"(seed={seed})"
    )
    return folds


def collect_all_keys(split_data):
    """Extract the union of all video keys from the original split JSON.

    The JSON can be a list-of-dicts (one per split) or a flat dict with a
    single set of keys.  In both cases we gather every unique key.

    Args:
        split_data: Parsed JSON content (list or dict).

    Returns:
        list[str]: Sorted list of all unique video keys.
    """
    keys = set()
    if isinstance(split_data, list):
        for entry in split_data:
            for k in ("train_keys", "test_keys", "val_keys"):
                keys.update(entry.get(k, []))
    else:
        for k in ("train_keys", "test_keys", "val_keys"):
            keys.update(split_data.get(k, []))
    return sorted(keys)


# ---------------------------------------------------------------------------
# Core inference  (unchanged)
# ---------------------------------------------------------------------------

def run_inference(model, data_path, keys, eval_method, save_summary,
                  dataset, verbose=False):
    """Run inference for a single model checkpoint over all test videos.

    Returns:
        mean_fscore, mean_kendall, mean_spearman,
        video_summaries, [video_names if SumMe]
    """
    model.eval()

    video_fscores   = []
    video_kendalls  = []
    video_spearmans = []
    video_summaries = {}
    video_names     = {}
    summe = (dataset.lower() == 'summe')

    for video in keys:
        if summe:
            try:
                if int(video.split('_')[1]) > 25:
                    continue
            except (IndexError, ValueError):
                pass

        frame_features, user_summary, sb, n_frames, positions, vname = \
            load_video_data(dataset, data_path, video)

        with torch.no_grad():
            scores, _ = model(frame_features)
            scores = scores.squeeze(0).cpu().numpy().tolist()

        summary = generate_summary([sb], [scores], [n_frames], [positions])[0]
        f_score = evaluate_summary(summary, user_summary, eval_method)

        frame_init_scores = np.array(scores)
        frame_scores      = np.zeros(n_frames, dtype=float)
        pos = positions.astype(int)
        if pos[-1] != n_frames:
            pos = np.concatenate([pos, [n_frames]])
        for i in range(len(pos) - 1):
            frame_scores[pos[i]:pos[i + 1]] = frame_init_scores[i]

        gt_importance = (
            user_summary.mean(axis=0) if user_summary.ndim > 1 else user_summary
        )

        if frame_scores.shape[0] != gt_importance.shape[0]:
            logging.warning(
                f"Shape mismatch for {video}: "
                f"pred={frame_scores.shape[0]}, gt={gt_importance.shape[0]}"
                " — skipping correlations"
            )
            ktau, spr = float('nan'), float('nan')
        else:
            ktau, _ = kendalltau(frame_scores, gt_importance)
            spr,  _ = spearmanr(frame_scores,  gt_importance)

        video_fscores.append(f_score)
        video_kendalls.append(ktau)
        video_spearmans.append(spr)
        video_summaries[video] = summary

        if summe:
            video_names[video] = vname

        if verbose:
            logging.info(
                f"  {video} ({vname}): F1={f_score:.2f}%  τ={ktau:.4f}  ρ={spr:.4f}"
            )

        if save_summary:
            out   = {str(i): int(v) for i, v in enumerate(summary)}
            fname = f"{video}_summary.json"
            with open(fname, "w") as fp:
                json.dump(out, fp, indent=4)
            print(f"Summary saved → {fname}")

    mean_fscore   = float(np.nanmean(video_fscores))
    mean_kendall  = float(np.nanmean(video_kendalls))
    mean_spearman = float(np.nanmean(video_spearmans))

    if summe:
        return mean_fscore, mean_kendall, mean_spearman, video_summaries, video_names
    return mean_fscore, mean_kendall, mean_spearman, video_summaries


# ---------------------------------------------------------------------------
# Parallel full-scan worker  (unchanged except docstring reference to CV)
# ---------------------------------------------------------------------------

def _scan_split_worker(args):
    """Evaluate all epoch checkpoints for a single CV fold.

    Designed to run inside a separate process via ProcessPoolExecutor.
    All arguments are plain Python objects (pickle-safe) — no live model or
    tensor objects are passed across the process boundary.

    Args:
        args: tuple of
            (split_id, model_path, epoch_files, dataset_path,
             test_keys, eval_metric, dataset, model_kwargs, verbose)

    Returns:
        (split_id, best_epoch, results_dict)
        where results_dict maps epoch_num → (fscore, kendall, spearman).
    """
    (split_id, model_path, epoch_files,
     dataset_path, test_keys,
     eval_metric, dataset, model_kwargs, verbose) = args

    results = {}
    for fname in epoch_files:
        epoch_num = int(re.findall(r'\d+', fname)[0])

        model = xLSTM(**model_kwargs)
        model.load_state_dict(
            torch.load(join(model_path, fname), map_location='cpu')
        )

        fs, kt, sp, *_ = run_inference(
            model, dataset_path, test_keys,
            eval_metric, save_summary=False,
            dataset=dataset, verbose=verbose,
        )
        results[epoch_num] = (fs, kt, sp)

    best_epoch = max(results, key=lambda e: results[e][0])
    return split_id, best_epoch, results


def _run_full_scan_parallel(split_ids, split_configs, n_workers):
    """Run full epoch scan for all CV folds in parallel."""
    n_workers = min(n_workers, len(split_ids))
    print(f"Full scan: {n_workers} parallel worker(s) across {len(split_ids)} folds\n")

    output = {}

    with ProcessPoolExecutor(max_workers=n_workers) as executor:
        futures = {
            executor.submit(_scan_split_worker, split_configs[s]): s
            for s in split_ids
        }

        for future in as_completed(futures):
            split_id = futures[future]
            try:
                sid, best_epoch, results = future.result()
                output[sid] = (best_epoch, results)
                best_fs = results[best_epoch][0]
                print(
                    f"  Fold {sid} done — best epoch: {best_epoch} "
                    f"(F1={best_fs:.2f}%)"
                )
            except Exception as exc:
                logging.error(f"  Fold {split_id} failed: {exc}")

    return output


# ---------------------------------------------------------------------------
# Output helpers  (unchanged)
# ---------------------------------------------------------------------------

def _print_results(split_ids, best_epochs, split_results,
                   best_avg_epoch, avg_fs, avg_ks, avg_ss):
    """Print a clean summary table to stdout."""
    sep = "-" * 62
    print(f"\n{sep}")
    print(
        f"{'Fold':<8} {'Epoch':>6}  {'F1 (%)':>8}  "
        f"{'Kendall τ':>10}  {'Spearman ρ':>11}"
    )
    print(sep)
    for s in split_ids:
        if s not in best_epochs:
            continue
        ep         = best_epochs[s]
        fs, kt, sp = split_results[s]
        print(f"  {s:<6} {ep:>6}  {fs:>8.2f}  {kt:>10.4f}  {sp:>11.4f}")
    print(sep)
    print(
        f"  {'AVG':<6} {best_avg_epoch:>6}  "
        f"{avg_fs[best_avg_epoch]:>8.2f}  "
        f"{avg_ks[best_avg_epoch]:>10.4f}  "
        f"{avg_ss[best_avg_epoch]:>11.4f}"
    )
    print(f"{sep}\n")


def _save_xlsx(split_ids, all_epoch_results, dataset):
    """Save full per-epoch metrics to an xlsx file."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    all_epochs = sorted({
        ep
        for s in split_ids
        for ep in all_epoch_results.get(s, {})
    })

    rows       = {"Epoch": all_epochs}
    avg_fs, avg_ks, avg_ss = {}, {}, {}

    for ep in all_epochs:
        vf = [all_epoch_results[s][ep][0] for s in split_ids if ep in all_epoch_results.get(s, {})]
        vk = [all_epoch_results[s][ep][1] for s in split_ids if ep in all_epoch_results.get(s, {})]
        vs = [all_epoch_results[s][ep][2] for s in split_ids if ep in all_epoch_results.get(s, {})]
        avg_fs[ep] = float(np.nanmean(vf)) if vf else float('nan')
        avg_ks[ep] = float(np.nanmean(vk)) if vk else float('nan')
        avg_ss[ep] = float(np.nanmean(vs)) if vs else float('nan')

    for s in split_ids:
        er = all_epoch_results.get(s, {})
        rows[f"F-score Fold {s}"]  = [er.get(ep, (None,))[0]           for ep in all_epochs]
        rows[f"Kendall Fold {s}"]  = [er.get(ep, (None, None))[1]       for ep in all_epochs]
        rows[f"Spearman Fold {s}"] = [er.get(ep, (None, None, None))[2] for ep in all_epochs]
    rows["Avg F-score"]  = [avg_fs[ep] for ep in all_epochs]
    rows["Avg Kendall"]  = [avg_ks[ep] for ep in all_epochs]
    rows["Avg Spearman"] = [avg_ss[ep] for ep in all_epochs]

    df = pd.DataFrame(rows).set_index("Epoch")

    tuples = []
    for s in split_ids:
        for m in ("F-score", "Kendall", "Spearman"):
            tuples.append((f"Fold {s}", m))
    for m in ("F-score", "Kendall", "Spearman"):
        tuples.append(("Average", m))
    df.columns = pd.MultiIndex.from_tuples(tuples)

    xlsx_path = f"{dataset}_cv_epoch_metrics.xlsx"
    df.to_excel(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.merge_cells("A1:A2")
    cell       = ws["A1"]
    cell.value = "Epoch"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.delete_rows(3, 1)
    wb.save(xlsx_path)

    print(f"Full epoch metrics saved → {xlsx_path}")


# ---------------------------------------------------------------------------
# Main  (adapted for cross-validation)
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Run inference with k-fold cross-validation and report "
            "best-epoch results per fold."
        )
    )
    parser.add_argument("--dataset",       type=str,   default='SumMe',
                        help="Dataset [SumMe | TVSum | MrHiSum]")
    parser.add_argument("--model_version", type=str,   default='',
                        help="Model version suffix, e.g. 'v2'")
    parser.add_argument("--verbose",       type=int,   default=0,
                        help="Per-video log (0=off, 1=on)")
    parser.add_argument("--save_summary",  type=int,   default=0,
                        help="Export binary summary JSON per video (0=off, 1=on)")
    parser.add_argument("--save_results",  type=int,   default=0,
                        help="Full epoch scan + save xlsx (0=off, 1=on)")
    parser.add_argument("--workers",       type=int,   default=5,
                        help="Parallel worker processes for full scan "
                             "(--save_results=1 only). Default=5. "
                             "Set to 1 to disable parallelism.")
    parser.add_argument("--hidden_dim",    type=int,   default=512)
    parser.add_argument("--num_layers",    type=int,   default=2)
    parser.add_argument("--dropout",       type=float, default=0.5,
                        help="Must match the value used during training.")
    # ---- NEW: cross-validation arguments ----
    parser.add_argument("--n_folds",       type=int,   default=5,
                        help="Number of cross-validation folds (default 5).")
    parser.add_argument("--cv_seed",       type=int,   default=42,
                        help="Random seed for fold shuffle (default 42).")

    args = vars(parser.parse_args())

    dataset       = args["dataset"]
    model_version = args["model_version"]
    verbose       = bool(args["verbose"])
    save_summary  = bool(args["save_summary"])
    save_results  = bool(args["save_results"])
    n_workers     = args["workers"]
    n_folds       = args["n_folds"]
    cv_seed       = args["cv_seed"]

    eval_metric = 'avg' if dataset.lower() == 'tvsum' else 'max'

    model_kwargs = dict(
        input_size=1024,
        output_size=1024,
        num_segments=4,
        hidden_dim=args["hidden_dim"],
        num_layers=args["num_layers"],
        dropout=args["dropout"],
    )

    paths        = get_paths(dataset)
    dataset_path = paths['dataset']
    split_file   = paths['split']

    with open(split_file) as fp:
        split_data = json.load(fp)

    # ---- Gather all keys and build CV folds on-the-fly ----
    all_keys   = collect_all_keys(split_data)
    cv_folds   = build_cv_folds(all_keys, n_folds=n_folds, seed=cv_seed)
    fold_ids   = list(range(len(cv_folds)))

    print(
        f"\nDataset: {dataset}  |  eval: {eval_metric}  |  "
        f"CV folds: {n_folds}  |  seed: {cv_seed}  |  "
        f"total videos: {len(all_keys)}"
    )

    # -----------------------------------------------------------------------
    # Full scan path (--save_results 1): parallel processing across folds
    # -----------------------------------------------------------------------
    if save_results:
        split_configs = {}
        for fold_id in fold_ids:
            model_path = (
                f"Summaries/xLSTM/{dataset}{model_version}/models/fold{fold_id}"
            )
            test_keys   = cv_folds[fold_id]["test_keys"]
            epoch_files = _find_epoch_files(model_path)

            if not epoch_files:
                logging.warning(
                    f"No epoch files in {model_path} — skipping fold {fold_id}"
                )
                continue

            split_configs[fold_id] = (
                fold_id, model_path, epoch_files,
                dataset_path, test_keys,
                eval_metric, dataset, model_kwargs, verbose,
            )

        if not split_configs:
            print("No valid folds found. Check model paths.")
            return

        scan_output = _run_full_scan_parallel(
            list(split_configs.keys()), split_configs, n_workers
        )

        best_epochs       = {}
        split_results     = {}
        all_epoch_results = {}

        for sid, (best_epoch, results) in scan_output.items():
            best_epochs[sid]       = best_epoch
            split_results[sid]     = results[best_epoch]
            all_epoch_results[sid] = results

    # -----------------------------------------------------------------------
    # Fast path (--save_results 0): one inference per fold
    # -----------------------------------------------------------------------
    else:
        best_epochs   = {}
        split_results = {}

        for fold_id in fold_ids:
            model_path = (
                f"Summaries/xLSTM/{dataset}{model_version}/models/fold{fold_id}"
            )
            test_keys   = cv_folds[fold_id]["test_keys"]
            epoch_files = _find_epoch_files(model_path)

            if not epoch_files:
                logging.warning(
                    f"No epoch files in {model_path} — skipping fold {fold_id}"
                )
                continue

            best_epoch = _load_best_epoch_from_fscores(model_path)

            if best_epoch is not None:
                best_pkl = join(model_path, 'best_model.pkl')
                fname    = (
                    'best_model.pkl'
                    if os.path.exists(best_pkl)
                    else f'epoch-{best_epoch}.pkl'
                )
                print(f"Fold {fold_id}: epoch {best_epoch} (from f_scores.txt)")
            else:
                fname      = epoch_files[-1]
                best_epoch = int(re.findall(r'\d+', fname)[0])
                print(
                    f"Fold {fold_id}: f_scores.txt not found — "
                    f"using last epoch ({best_epoch})"
                )

            model = xLSTM(**model_kwargs)
            model.load_state_dict(
                torch.load(join(model_path, fname), map_location='cpu')
            )
            fs, kt, sp, *_ = run_inference(
                model, dataset_path, test_keys,
                eval_metric, save_summary, dataset, verbose,
            )
            best_epochs[fold_id]   = best_epoch
            split_results[fold_id] = (fs, kt, sp)

    # -----------------------------------------------------------------------
    # Print consolidated results
    # -----------------------------------------------------------------------
    if not split_results:
        print("No results collected — check model paths and fold directories.")
        return

    valid_folds = list(split_results.keys())
    all_f = [split_results[s][0] for s in valid_folds]
    all_k = [split_results[s][1] for s in valid_folds]
    all_s = [split_results[s][2] for s in valid_folds]

    best_avg_epoch = best_epochs[max(valid_folds, key=lambda s: split_results[s][0])]
    avg_fs = {best_avg_epoch: float(np.nanmean(all_f))}
    avg_ks = {best_avg_epoch: float(np.nanmean(all_k))}
    avg_ss = {best_avg_epoch: float(np.nanmean(all_s))}

    # ---- NEW: print std alongside mean for a richer CV report ----
    std_f = float(np.nanstd(all_f))
    std_k = float(np.nanstd(all_k))
    std_s = float(np.nanstd(all_s))

    _print_results(
        valid_folds, best_epochs, split_results,
        best_avg_epoch, avg_fs, avg_ks, avg_ss,
    )

    print(
        f"  Std dev across folds →  "
        f"F1: {std_f:.2f}%   τ: {std_k:.4f}   ρ: {std_s:.4f}\n"
    )

    if save_results:
        _save_xlsx(valid_folds, all_epoch_results, dataset)


if __name__ == "__main__":
    main()
